"""将业务规则文件标准化为可供规则草稿生成器使用的文本。"""

from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import load_workbook


SUPPORTED_RULE_FILE_EXTENSIONS = {".txt", ".xlsx", ".xlsm"}


def read_rule_file(filename: str, content: bytes) -> str:
    """读取 TXT 或“字段 / 规则描述”格式的 Excel 规则文件。"""
    extension = _extension(filename)
    if extension == ".txt":
        try:
            return content.decode("utf-8-sig").strip()
        except UnicodeDecodeError as error:
            raise ValueError("TXT 文件必须使用 UTF-8 编码") from error
    if extension not in SUPPORTED_RULE_FILE_EXTENSIONS:
        raise ValueError("仅支持 TXT、XLSX 或 XLSM 规则文件")

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as error:
        raise ValueError("Excel 文件无法读取，请确认文件未损坏且格式为 XLSX 或 XLSM") from error

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        columns = _find_rule_columns(rows)
        if columns is None:
            continue
        field_column, description_column, header_row = columns
        lines = _rule_lines(rows[header_row + 1 :], field_column, description_column)
        if lines:
            return "\n".join(lines)
    raise ValueError("未找到“录入ERP的字段”和“规则描述”两列，请使用示例表头")


def _extension(filename: str) -> str:
    dot = filename.rfind(".")
    return filename[dot:].lower() if dot >= 0 else ""


def _find_rule_columns(rows: Iterable[tuple]) -> tuple[int, int, int] | None:
    for row_index, row in enumerate(rows):
        normalized = [_text(value).replace(" ", "") for value in row]
        field_column = next((index for index, value in enumerate(normalized) if "录入erp的字段" in value.lower()), None)
        description_column = next((index for index, value in enumerate(normalized) if "规则描述" in value), None)
        if field_column is not None and description_column is not None:
            return field_column, description_column, row_index
    return None


def _rule_lines(rows: Iterable[tuple], field_column: int, description_column: int) -> list[str]:
    lines: list[str] = []
    for row in rows:
        field = _text(row[field_column] if field_column < len(row) else None)
        description = _text(row[description_column] if description_column < len(row) else None)
        if field and description:
            lines.append(f"{field}：{description}")
    return lines


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()
