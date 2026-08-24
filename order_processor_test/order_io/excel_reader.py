"""Excel 读取器。"""

from pathlib import Path
from typing import Any, Dict, List
from zipfile import ZipFile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


class ExcelReader:
    """读取首个工作表，并将每一行转换为字典。"""

    @staticmethod
    def read(file_path: str) -> List[Dict[str, Any]]:
        path = Path(file_path)
        if not path.is_file():
            raise FileNotFoundError(f"输入 Excel 不存在: {path}")

        raw_values = ExcelReader._raw_cell_values(path)
        workbook = load_workbook(path, data_only=True, read_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows()
            header_row = next(rows, None)
            if not header_row:
                return []
            headers = [str(cell.value).strip() if cell.value is not None else f"col_{index}"
                       for index, cell in enumerate(header_row)]
            records: List[Dict[str, Any]] = []
            for row in rows:
                if not any(cell.value is not None for cell in row):
                    continue
                record: Dict[str, Any] = {}
                for index, cell in enumerate(row):
                    if index >= len(headers):
                        continue
                    header = headers[index]
                    value = ExcelReader._cell_value(cell, raw_values)
                    # 客户表中常见重复表头（如“当前日期”）；后续空列不能覆盖前面有效值。
                    if header not in record or (record[header] in (None, "") and value not in (None, "")):
                        record[header] = value
                records.append(record)
            return records
        finally:
            workbook.close()

    @staticmethod
    def _cell_value(cell: Any, raw_values: Dict[str, str]) -> Any:
        """保留 Excel 中被误标为日期的八位数字，例如 20260821。"""
        if cell.data_type == "e" and cell.value == "#VALUE!" and cell.coordinate in raw_values:
            return raw_values[cell.coordinate]
        return cell.value

    @staticmethod
    def _raw_cell_values(path: Path) -> Dict[str, str]:
        """从第一个工作表 XML 读取原始值，补救 openpyxl 对非法日期序列的转换。"""
        namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        with ZipFile(path) as archive:
            root = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))
        result: Dict[str, str] = {}
        for cell in root.findall(".//main:c", namespace):
            value = cell.find("main:v", namespace)
            if value is not None and value.text is not None:
                result[cell.attrib["r"]] = value.text
        return result
